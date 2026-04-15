from docx import Document
import io
from io import BytesIO
import re
import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

bulan_pattern = re.compile(
    r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Mei|Agu|Okt|Des)",
    re.IGNORECASE,
)


def _clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def _normalize_date(value):
    text = _clean_text(value)
    return text.replace("\u2010", "-").replace("\u2013", "-")


def is_dark_fill(cell):
    try:
        shading = cell._element.xpath(".//w:shd")
        if shading:
            fill = shading[0].get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}fill")
            if fill and fill.lower() in ["000000", "1f1f1f", "222222", "2e2e2e", "444444"]:
                return True
    except Exception:
        pass
    return False


def detect_file_type(input_file):
    name = (input_file.filename or "").lower()
    if name.endswith(".docx"):
        return "docx"
    if name.endswith(".pdf"):
        return "pdf"
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return "excel"
    return None


def is_supported_file(input_file):
    return detect_file_type(input_file) in {"docx", "pdf", "excel"}


def is_docx(input_file):
    return detect_file_type(input_file) == "docx"


def read_file(input_file):
    file_type = detect_file_type(input_file)
    if file_type == "docx":
        input_file.seek(0)
        return Document(io.BytesIO(input_file.read()))
    raise ValueError("Format file tidak didukung.")


def _extract_date_columns(header_cells):
    date_cols = []
    for idx, raw in enumerate(header_cells):
        cell = _normalize_date(raw)
        if not cell or not bulan_pattern.search(cell):
            continue
        if date_cols and idx - date_cols[-1][0] < 4:
            continue
        date_cols.append((idx, cell))
    return date_cols


def _detect_name_col(header_cells):
    for idx, cell in enumerate(header_cells):
        if "nama" in _clean_text(cell).lower():
            return idx
    return 1


def _build_schedules_from_rows(rows):
    schedules = []
    name_col = 1
    date_cols = []

    for row in rows:
        lowered = [_clean_text(c).lower() for c in row]
        has_nama = any("nama" in c for c in lowered)
        has_npm = any("npm" in c for c in lowered)
        found_dates = _extract_date_columns(row)

        if has_nama and (has_npm or found_dates):
            name_col = _detect_name_col(row)
            date_cols = found_dates
            break

        if not date_cols and found_dates:
            date_cols = found_dates

    if not date_cols:
        return schedules

    for row in rows:
        if len(row) <= name_col:
            continue

        nama = _clean_text(row[name_col]).lower()
        if not nama or nama == "nama":
            continue

        jadwal = []
        any_slot_filled = False

        for start_col, tanggal in date_cols:
            sesi = []
            for offset in range(4):
                col = start_col + offset
                val = _clean_text(row[col]) if col < len(row) else ""
                if val:
                    any_slot_filled = True
                sesi.append(val if val else "-")
            jadwal.append({"tanggal": tanggal, "sesi": sesi})

        if jadwal and any_slot_filled:
            schedules.append({"nama": nama, "jadwal": jadwal})

    return schedules


def _all_schedules_docx(input_docx):
    schedules = []
    for table in input_docx.tables:
        rows = []
        for row in table.rows:
            cells = []
            for c in row.cells:
                text = _clean_text(c.text)
                if is_dark_fill(c) and not text:
                    cells.append("")
                else:
                    cells.append(text)
            if any(cells):
                rows.append(cells)

        if rows:
            schedules.extend(_build_schedules_from_rows(rows))

    return schedules


def _all_schedules_pdf(input_file):
    schedules = []
    input_file.seek(0)
    file_bytes = input_file.read()

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                rows = []
                for row in table:
                    if not row:
                        continue
                    cells = [_clean_text(c) for c in row]
                    if any(cells):
                        rows.append(cells)

                if rows:
                    schedules.extend(_build_schedules_from_rows(rows))

    return schedules


def _all_schedules_excel(input_file):
    schedules = []
    input_file.seek(0)
    file_bytes = input_file.read()

    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [_clean_text(c) for c in row]
            if any(cells):
                rows.append(cells)

        if rows:
            schedules.extend(_build_schedules_from_rows(rows))

    return schedules


def all_schedules(input_file):
    file_type = detect_file_type(input_file)
    if file_type == "docx":
        return _all_schedules_docx(read_file(input_file))
    if file_type == "pdf":
        return _all_schedules_pdf(input_file)
    if file_type == "excel":
        return _all_schedules_excel(input_file)
    return []


def is_valid_jadwal(file):
    try:
        return len(all_schedules(file)) > 0
    except Exception:
        return False


def find_asisten(schedules, nama_asisten):
    return any(a["nama"].lower() == nama_asisten.lower() for a in schedules)


def find_patners(all_schedules, target_name):
    target = next((a for a in all_schedules if a["nama"].lower() == target_name.lower()), None)
    if not target:
        return []

    hasil = []
    for j in target["jadwal"]:
        tanggal = j["tanggal"]
        for sesi_idx, ruang in enumerate(j["sesi"]):
            if ruang == "-" or not ruang:
                continue

            patners = []
            for other in all_schedules:
                if other["nama"].lower() == target_name.lower():
                    continue
                for oj in other["jadwal"]:
                    if oj["tanggal"] == tanggal and oj["sesi"][sesi_idx] == ruang:
                        patners.append(other["nama"])
                        break

            hasil.append({
                "ruangan": f"{tanggal} | Sesi {sesi_idx+1} | {ruang}",
                "patners": patners if patners else ["-"],
            })

    return hasil


def generate_excel(jadwal, patners, nama):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal Ngawas {nama[0].title()}"

    fill_header = PatternFill("solid", fgColor="DCE6F1")
    fill_sesi = PatternFill("solid", fgColor="FCE4D6")
    border_medium = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )
    center = Alignment(horizontal="center", vertical="center")

    total_tanggal = len(jadwal)
    total_kolom = total_tanggal + 1
    last_col_letter = get_column_letter(total_kolom)

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"Jadwal Mengawas - {nama.title()}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    ws["A2"] = "Sesi / Tanggal"
    ws["A2"].fill = fill_header
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = center
    ws["A2"].border = border_medium

    for i, j in enumerate(jadwal, start=2):
        c = ws.cell(row=2, column=i, value=j["tanggal"])
        c.fill = fill_header
        c.font = Font(bold=True)
        c.alignment = center
        c.border = border_medium

    for s_idx in range(4):
        sesi_cell = ws.cell(row=3 + s_idx, column=1, value=f"{s_idx + 1}")
        sesi_cell.fill = fill_sesi
        sesi_cell.border = border_medium
        sesi_cell.alignment = center
        for c_idx, j in enumerate(jadwal, start=2):
            val = j["sesi"][s_idx]
            cell = ws.cell(row=3 + s_idx, column=c_idx, value=val)
            cell.border = border_medium
            cell.alignment = center

    start_row = 9
    ws.merge_cells(f"A{start_row}:D{start_row}")
    ws[f"A{start_row}"] = "Daftar Patners"
    ws[f"A{start_row}"].font = Font(bold=True, size=14)
    ws[f"A{start_row}"].alignment = center

    ws[f"A{start_row+1}"] = "Ruangan"
    ws.merge_cells(f"B{start_row+1}:D{start_row+1}")
    ws[f"B{start_row+1}"] = "Patners"
    for col in ["A", "B", "C", "D"]:
        cell = ws[f"{col}{start_row+1}"]
        cell.fill = fill_header
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border_medium

    for i, p in enumerate(patners, start=start_row + 2):
        ws[f"A{i}"] = p["ruangan"]
        ws[f"A{i}"].fill = fill_sesi
        ws.merge_cells(f"B{i}:D{i}")
        ws[f"B{i}"] = ", ".join(p["patners"]) if p["patners"][0] != "-" else "-"
        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{i}"]
            cell.border = border_medium
            cell.alignment = center

    for col in range(1, total_kolom + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 20

    ws.column_dimensions["A"].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
